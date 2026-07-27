import time, sys, os, io, json, zipfile, csv, threading, math, getpass
import requests, pyautogui, ctypes, pandas as pd, openpyxl
import msoffcrypto
from datetime import datetime, date
from openpyxl import load_workbook


# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN GLOBAL
# ═══════════════════════════════════════════════════════════════

MAGPY_HOST            = "http://localhost:6000"
ACQUISITION_PERIOD_MS = 700
# Sin sleep en el bucle de measurement — la latencia HTTP (~100-200ms) ya actúa
# como cadencia natural. Así nunca perdemos frames intermedios.
POLL_INTERVAL_S       = 0.0

DURACION_6MIN  =  6 * 60
DURACION_30MIN = 30 * 60
DURACION_2MIN  =  2 * 60

LOCATIONS       = ["ProbeCenter", "SensorMax", "TipCenter", "TipMax"]
QUANTITY_FILTRO = "IncidentMagneticField"

# Consola reducida: solo mensajes realmente necesarios para el operador.
# Cambia a "NORMAL" si algún día quieres volver a ver todos los mensajes técnicos.
LOG_MODE = "ESSENTIAL"

# Reintentos de arranque de adquisición. La posición actual no se salta
# automáticamente si falla el inicio de medida.
MAX_REINTENTOS_INICIO_MEDIDA = 8
ESPERA_REINTENTO_INICIO_S    = 2


# Formato de exportación MAGPy V3.4 para Measurements.csv.
# No se pulsa el botón SAVE DATA. Se transforma la lectura de la API al
# formato largo documentado por MAGPy 3.4.
API_LOCATIONS_EXPORT = ["probe-center", "sensor-max", "tip-center", "tip-max"]
LOCATION_API_TO_EXPORT = {
    "probe-center": "ProbeCenter",
    "sensor-max":   "SensorMax",
    "tip-center":   "TipCenter",
    "tip-max":      "TipMax",
}
MEASUREMENTS_TIDY_COLUMNS = [
    "timestamp", "frame_source", "quantity", "location", "sensor_index",
    "frequency", "x", "y", "z", "total"
]

MS_6MIN         =  6 * 60 * 1000
MS_30MIN        = 30 * 60 * 1000
MS_2MIN         =  2 * 60 * 1000

CONFIG_REGION = {
    "CE": {
        "nombre":              "CE (Europa - ICNIRP 2010/2020)",
        "standard":            "ICNIRP 2010/2020",
        "environment":         "General public",
        "health_effect":       "Combined",
        "peak_rms_inicial":    "RMS",
        "sheet":               "EN Results",
        "fases": [
            {"etiqueta": "6 min RMS", "duracion_s": DURACION_6MIN,
             "peak_rms": "RMS", "ms_ventana": MS_6MIN},
        ],
    },
    "FCC": {
        "nombre":              "FCC (Estados Unidos)",
        "standard":            "FCC",
        "environment":         "General public",
        "health_effect":       "Combined",
        "peak_rms_inicial":    "RMS",
        "sheet":               "FCC Results",
        "fases": [
            {"etiqueta": "30 min RMS", "duracion_s": DURACION_30MIN,
             "peak_rms": "RMS", "ms_ventana": MS_30MIN},
        ],
    },
    "ISED": {
        "nombre":              "ISED / HC Code 6 (Canada)",
        "standard":            "HC Code 6",
        "environment":         "General public",
        "health_effect":       "Combined",
        "peak_rms_inicial":    "RMS",
        "sheet":               "ISED Results",
        "fases": [
            {"etiqueta": "6 min RMS",  "duracion_s": DURACION_6MIN,
             "peak_rms": "RMS",  "ms_ventana": MS_6MIN},
            {"etiqueta": "2 min Pico", "duracion_s": DURACION_2MIN,
             "peak_rms": "Peak", "ms_ventana": MS_2MIN},
        ],
    },
}

EXCEL_COL_MAP = {
    "CE": [
        {   "technology": 2,  "wpt": 3,      "battery": 4,
            "testside": 5,    "distance": 6,
            "freq_e": 7,      "efield": 8,   "date_e": 12,
            "freq_h": 16,     "hfield": 17,  "date_h": 21,
            "first_data_row": 9,
        }
    ],
    "FCC": [
        {   "technology": 2,  "wpt": 3,      "battery": 4,
            "testside": 5,    "distance": 6,
            "freq_e": 7,      "efield": 8,   "date_e": 12,
            "freq_h": 16,     "hfield": 17,  "date_h": 21,
            "first_data_row": 9,
        },
        {   "technology": 26, "wpt": 27,     "battery": 28,
            "testside": 29,   "distance": 30,
            "freq_e": 31,     "efield": 32,  "date_e": 36,
            "freq_h": 40,     "hfield": 41,  "date_h": 45,
            "first_data_row": 9,
        },
    ],
    "ISED": [
        {   "technology": 2,  "wpt": 3,      "battery": 4,
            "testside": 5,    "distance": 6,
            "freq_e": 7,      "efield": 8,   "date_e": 12,
            "freq_h": 16,     "hfield": 17,  "date_h": 21,
            "first_data_row": 9,
        },
        {   "technology": 27, "wpt": 28,     "battery": 29,
            "testside": 30,   "distance": 31,
            "freq_e": 32,     "efield": 33,  "date_e": 37,
            "freq_h": 41,     "hfield": 42,  "date_h": 46,
            "first_data_row": 9,
        },
        {   "technology": 52, "wpt": 53,     "battery": 54,
            "testside": 55,   "distance": 56,
            "freq_e": 57,     "efield": 58,  "date_e": 62,
            "freq_h": 66,     "hfield": 67,  "date_h": 71,
            "first_data_row": 9,
        },
    ],
}


# ═══════════════════════════════════════════════════════════════
# UTILIDADES DE CONSOLA
# ═══════════════════════════════════════════════════════════════

def beep(n=1):
    for _ in range(n):
        ctypes.windll.kernel32.Beep(1000, 500)
        time.sleep(0.2)

def _mensaje_esencial(texto, tipo):
    """Filtro central para reducir ruido en consola sin tocar la lógica de medida."""
    if LOG_MODE.upper() != "ESSENTIAL":
        return True

    texto_l = str(texto).lower()

    # Estos mensajes son útiles para depuración, pero demasiado ruidosos para operación normal.
    ocultar = (
        "endpoint histórico encontrado",
        "no hay endpoint histórico",
        "config raw recibida",
        "campo '",
        "configuración obtenida desde endpoint",
        "configuracion obtenida desde endpoint",
        "empaquetando",
        "captura guardada",
        "promediar_medidas:",
        "excel '",
        "api:",
    )
    if any(p in texto_l for p in ocultar):
        return False

    # Errores y avisos reales se mantienen, salvo los ocultados arriba.
    if tipo in ("ERROR", "AVISO"):
        return True

    # Mensajes OK/INFO/ESPERA que sí necesita ver el operador.
    esenciales = (
        "verificando conexion",
        "conectado",
        "sin conexion",
        "validando configuracion",
        "configuracion validada",
        "configuración validada",
        "medida iniciada",
        "medida detenida",
        "magpy ya estaba midiendo",
        "reintent",
        "no se pudo",
        "zip guardado",
        "excel guardado",
        "csv de resumen",
        "promedios calculados",
        "excel actualizado",
        "posicion '",
        "posición '",
        "sesion completada",
        "sesión completada",
    )
    return any(p in texto_l for p in esenciales)

def msg(texto, tipo="INFO"):
    if not _mensaje_esencial(texto, tipo):
        return
    ts = datetime.now().strftime("%H:%M:%S")
    iconos = {"INFO": "i", "OK": "OK", "AVISO": "!", "ERROR": "X", "ESPERA": "..."}
    print(f"[{ts}] [{iconos.get(tipo,'?')}]  {texto}")

def sep(titulo=""):
    linea = "-" * 62
    print(f"\n{linea}\n  {titulo}\n{linea}" if titulo else linea)

def cuadro(titulo, lineas):
    ancho = max(len(l) for l in [titulo] + lineas) + 4
    print()
    print("+" + "=" * ancho + "+")
    print("|" + f"  {titulo}".ljust(ancho) + "|")
    print("+" + "=" * ancho + "+")
    for l in lineas:
        print("|" + f"  {l}".ljust(ancho) + "|")
    print("+" + "=" * ancho + "+")
    print()

def enter(texto="Presiona ENTER para continuar..."):
    input(f"\n  --> {texto}")


class ReiniciarConfiguracion(Exception):
    """Señal interna para volver a pedir todos los parámetros de la sesión."""
    pass


class VolverAtrasConfiguracion(Exception):
    """Señal interna para volver un paso atrás en la configuración."""
    pass


class CancelarSesion(Exception):
    """Señal interna para cancelar la sesión antes de empezar a medir."""
    pass


def input_config(prompt, permitir_vacio=False):
    """Entrada de parámetros con comandos de control comunes.

    Comandos disponibles durante la configuración inicial:
      B -> volver al parámetro anterior
      R -> reiniciar la selección completa de parámetros
      Q -> cancelar/salir del programa
    """
    while True:
        valor = input(prompt).strip().strip('"')
        comando = valor.upper()
        if comando == "B":
            raise VolverAtrasConfiguracion
        if comando == "R":
            raise ReiniciarConfiguracion
        if comando == "Q":
            raise CancelarSesion
        if valor or permitir_vacio:
            return valor
        print("  X Introduce un valor valido, o escribe B para atras / R para reiniciar / Q para cancelar.")


def _fmt_actual(valor):
    if valor is None:
        return "No configurado"
    if isinstance(valor, list):
        return " ".join(str(v) for v in valor)
    return str(valor)


def _prompt_actual(prompt, actual):
    if actual is None:
        return prompt
    prompt_base = prompt[:-2] if prompt.endswith(": ") else prompt
    return f"{prompt_base} [actual: {_fmt_actual(actual)}]: "


def mostrar_ayuda_parametros():
    print("\n  Comandos durante la configuracion:")
    print("    B  --> volver al parametro anterior")
    print("    R  --> reiniciar la seleccion de parametros desde cero")
    print("    Q  --> cancelar y salir")
    print("    ENTER sobre un campo con [actual: ...] conserva ese valor")

def barra_progreso(segundos_total, etiqueta="Midiendo"):
    """
    Barra de progreso con actualización visual cada 1%.
    Mantiene la duración real de la fase y evita acumulación de deriva usando time.monotonic().
    """
    segundos_total = max(1, int(segundos_total))
    inicio = time.monotonic()

    for pct in range(0, 101):
        transcurrido = min(time.monotonic() - inicio, segundos_total)
        restante = max(0, int(round(segundos_total - transcurrido)))
        mr, sr = divmod(restante, 60)

        barra = "#" * int(pct / 2) + "-" * (50 - int(pct / 2))
        print(f"\r  {etiqueta}: [{barra}] {pct:3d}%  Restante: {mr:02d}:{sr:02d}   ",
              end="", flush=True)

        if pct < 100:
            siguiente = inicio + segundos_total * ((pct + 1) / 100.0)
            espera = max(0.0, siguiente - time.monotonic())
            time.sleep(espera)

    print()


# ═══════════════════════════════════════════════════════════════
# API MAGPY3
# ═══════════════════════════════════════════════════════════════

# Sesión persistente con keep-alive: reutiliza la conexión TCP
# y reduce la latencia por request de ~400ms a ~50ms en localhost.
_session = requests.Session()
_session.headers.update({"Connection": "keep-alive"})

# Contraseña del Excel protegido: se pide una vez por consola (input oculto)
# y se cachea SOLO en memoria para el resto de la sesión. Nunca se escribe a disco.
_EXCEL_PASSWORD = None

def api_get(endpoint):
    try:
        r = _session.get(f"{MAGPY_HOST}/{endpoint}", timeout=5)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None

def api_post(endpoint, datos):
    try:
        r = _session.post(f"{MAGPY_HOST}/{endpoint}", json=datos, timeout=5)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None

def verificar_conexion():
    msg("Verificando conexion con MAGpy3...", "ESPERA")
    info = api_get("version")
    if info:
        msg(f"Conectado -- MAGpy3 version {info.get('application','?')}", "OK")
        return True
    msg("Sin conexion con MAGpy3", "ERROR")
    return False

def iniciar_medida(silencioso=False):
    r = api_get("acquisition/start")
    if r and r.get("active"):
        if not silencioso:
            msg("Medida iniciada", "OK")
        return True
    if not silencioso:
        msg("No se pudo iniciar la medida", "ERROR")
    return False

def adquisicion_activa():
    """Comprueba si MAGPy informa que la adquisición está activa."""
    status = api_get("acquisition/status")
    return bool(status and status.get("active"))

def reset_adquisicion_para_reintento():
    """Reset suave entre intentos: stop + clear GUI, sin mensajes extra."""
    try:
        api_get("acquisition/stop")
    except Exception:
        pass
    time.sleep(0.8)
    try:
        click_clear_gui()
    except Exception:
        pass
    time.sleep(0.8)

def iniciar_medida_con_reintentos(contexto="medida", max_reintentos=MAX_REINTENTOS_INICIO_MEDIDA):
    """
    Intenta iniciar la adquisición sin saltar la posición actual.
    Si falla el bloque de intentos automáticos, pide intervención y vuelve a intentar
    la MISMA posición/fase. Solo devuelve False si el operador decide abortar.
    """
    ciclo = 1
    while True:
        for intento in range(1, max_reintentos + 1):
            if iniciar_medida(silencioso=True):
                msg(f"Medida iniciada ({contexto})", "OK")
                return True

            time.sleep(ESPERA_REINTENTO_INICIO_S)

            if adquisicion_activa():
                msg(f"MAGPy ya estaba midiendo ({contexto}). Continuando.", "OK")
                return True

            # Cada 3 intentos hacemos un reset suave antes de volver a arrancar.
            if intento in (3, 6):
                msg(f"Reintentando {contexto}: reset suave de adquisicion ({intento}/{max_reintentos})", "AVISO")
                reset_adquisicion_para_reintento()
            elif intento in (1, max_reintentos):
                msg(f"Reintentando {contexto}: intento {intento}/{max_reintentos}", "AVISO")

        beep(2)
        cuadro(
            "NO SE HA PODIDO INICIAR LA MEDIDA",
            [f"Fase/posicion actual: {contexto}",
             "La posicion actual NO se va a saltar automaticamente.",
             "Revisa MAGPy, la sonda, el DUT y la conexion API.",
             "",
             "ENTER  --> reintentar la misma fase/posicion",
             "A      --> abortar la sesion"]
        )
        accion = input("  Opcion: ").strip().upper()
        if accion == "A":
            msg(f"Sesion abortada por el operador durante {contexto}.", "ERROR")
            return False
        ciclo += 1
        msg(f"Reintentando {contexto} de nuevo (bloque {ciclo}).", "AVISO")

def obtener_config_actual():
    """
    Consulta la configuración actual del software MAGPy.
    Prueba varios endpoints posibles porque la API puede variarlos entre versiones.
    """
    for endpoint in ("settings", "configuration", "config", "compliance/settings"):
        datos = api_get(endpoint)
        if datos:
            msg(f"Configuración obtenida desde endpoint '/{endpoint}'", "OK")
            return datos
    msg("Ningún endpoint de configuración respondió (settings/configuration/config/compliance/settings)", "AVISO")
    return None

def validar_configuracion_magpy(region_seleccionada):
    """
    Lee la config actual de MAGPy y la compara con la requerida.
    Muestra los campos raw que devuelve la API para facilitar el diagnóstico.
    """
    msg("Validando configuracion normativa en MAGpy3...", "ESPERA")
    config_actual = obtener_config_actual()

    if not config_actual:
        msg("No se pudo obtener la configuracion actual de MAGpy3 -- "
            "comprueba que la versión de MAGpy3 expone /settings o /configuration.", "AVISO")
        return True  # No bloqueamos la medida, pero avisamos

    # Log de los campos raw para diagnóstico futuro
    campos_raw = ", ".join(f"{k}={v}" for k, v in list(config_actual.items())[:8])
    msg(f"Config raw recibida: {campos_raw}", "INFO")

    cfg_esperada = CONFIG_REGION[region_seleccionada]

    # Mapeo ampliado: clave en nuestro dict -> posibles nombres en la API
    check_list = [
        ("standard",      ["standard", "norm", "normative_standard", "compliance_standard"]),
        ("environment",   ["environment", "env", "exposure_environment"]),
        ("health_effect", ["health_effect", "healthEffect", "health_effects", "effect"]),
    ]

    errores = []
    for dict_key, api_candidates in check_list:
        val_esperado = cfg_esperada.get(dict_key)
        val_actual   = None
        for candidate in api_candidates:
            if candidate in config_actual:
                val_actual = config_actual[candidate]
                break

        if val_actual is None:
            # Campo no encontrado en la respuesta: lo anotamos pero no bloqueamos
            msg(f"Campo '{dict_key}' no encontrado en la respuesta de la API "
                f"(buscado como: {api_candidates})", "AVISO")
            continue

        if str(val_actual).lower() != str(val_esperado).lower():
            errores.append(
                f"  * {dict_key.replace('_', ' ').title()}: "
                f"Actual='{val_actual}' | ESPERADO='{val_esperado}'"
            )

    if errores:
        cuadro("CONFIGURACION INCORRECTA",
               ["La configuracion en MAGpy3 NO coincide con el estandar requerido:", ""] + errores)
        return False

    msg("Configuracion validada correctamente.", "OK")
    return True

def parar_medida(reintentos=3, espera_s=2):
    """Detiene la adquisición con reintentos. Tolera que ya esté parada."""
    for intento in range(reintentos):
        r = api_get("acquisition/stop")
        # Éxito si la API confirma que ya no está activa
        if r is not None and not r.get("active", True):
            msg("Medida detenida", "OK")
            return True
        # Si r es None o active sigue en True, comprobar estado actual
        status = api_get("acquisition/status")
        if status and not status.get("active", True):
            msg("Medida ya estaba detenida (confirmado por status)", "OK")
            return True
        msg(f"Intento {intento+1}/{reintentos}: No se pudo detener la medida, reintentando...", "AVISO")
        time.sleep(espera_s)
    msg("No se pudo detener la medida tras varios intentos", "ERROR")
    return False

def configurar_adquisicion(frecuencia_khz):
    bw = round(frecuencia_khz * 0.20, 1)
    r  = api_post("acquisition/configure", {
        "period_ms": ACQUISITION_PERIOD_MS,
        "peak_search_kHz": {"frequency": frecuencia_khz, "bandwidth": bw},
    })
    if r:
        msg(f"API: {frecuencia_khz} kHz, BW={bw} kHz, periodo={ACQUISITION_PERIOD_MS} ms", "OK")
        return True
    return False



# ═══════════════════════════════════════════════════════════════
# TRANSFORMACIÓN API -> FORMATO MEASUREMENTS MAGPy 3.4
# ═══════════════════════════════════════════════════════════════

def normalizar_frame_source(source):
    """Convierte el campo opcional source de la API al nombre usado en el CSV MAGPy 3.4."""
    if source is None:
        return "Unknown"
    s = str(source).strip()
    compact = s.lower().replace("_", "").replace("-", "").replace(" ", "")
    if compact in ("periodic", "live"):
        return "Live"
    if compact in ("epeak", "electricfieldpeak"):
        return "ElectricFieldPeak"
    if compact in ("hpeak", "magneticfieldpeak"):
        return "MagneticFieldPeak"
    if s in ("ElectricFieldPeak", "MagneticFieldPeak", "Live", "Unknown"):
        return s
    return "Unknown"

def _to_float_or_none(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def _vector_total(x, y, z):
    vals = [_to_float_or_none(x), _to_float_or_none(y), _to_float_or_none(z)]
    if any(v is None for v in vals):
        return None
    return math.sqrt(vals[0] ** 2 + vals[1] ** 2 + vals[2] ** 2)

def _campo_xyz_total(campo, total_keys=("rms_total", "total")):
    """Extrae x/y/z/total de un campo de la API aceptando variantes de nombre."""
    if not isinstance(campo, dict):
        return None
    x = campo.get("rms_x", campo.get("x"))
    y = campo.get("rms_y", campo.get("y"))
    z = campo.get("rms_z", campo.get("z"))
    total = None
    for k in total_keys:
        if k in campo:
            total = campo.get(k)
            break
    if total is None:
        total = _vector_total(x, y, z)
    return x, y, z, total

def _extraer_sensor_index(datos, api_location):
    """
    MAGPy 3.4 exporta sensor_index. El endpoint /data/current/measurement documentado
    no garantiza este campo, pero algunas versiones pueden incluirlo con nombres internos.
    Si no está disponible, se deja 0 para no inventar un sensor.
    """
    if api_location in ("probe-center", "tip-center"):
        return 0

    candidatos_directos = [
        "sensor_index", "sensorIndex", "sensor_idx", "sensor",
        "sensor_num", "sensorNum", "sensor_number", "sensorNumber",
        "sensor_num_surface", "sensorNumSurface",
    ]
    for k in candidatos_directos:
        v = datos.get(k) if isinstance(datos, dict) else None
        try:
            if v is not None and int(v) >= 0:
                return int(v)
        except (TypeError, ValueError):
            pass

    for contenedor in ("h_field", "measurement", "meta", "metadata", "location"):
        sub = datos.get(contenedor) if isinstance(datos, dict) else None
        if isinstance(sub, dict):
            for k in candidatos_directos:
                v = sub.get(k)
                try:
                    if v is not None and int(v) >= 0:
                        return int(v)
                except (TypeError, ValueError):
                    pass

    return 0

def _fila_tidy(timestamp_abs, ts_ref, frame_source, quantity, location, sensor_index,
               frequency, x, y, z, total):
    ts = timestamp_abs
    if ts is not None and ts_ref is not None:
        try:
            ts = int(round(float(ts) - float(ts_ref)))
        except (TypeError, ValueError):
            pass
    return {
        "timestamp": ts,
        "frame_source": frame_source,
        "quantity": quantity,
        "location": location,
        "sensor_index": sensor_index,
        "frequency": frequency,
        "x": x,
        "y": y,
        "z": z,
        "total": total,
    }

def construir_filas_measurement_tidy(datos_por_location, ts_ref):
    """
    Construye filas tipo MAGPy 3.4:
    timestamp, frame_source, quantity, location, sensor_index, frequency, x, y, z, total

    Se usan únicamente valores devueltos por la API. No se pulsa SAVE DATA y no se
    crea el formato legacy ancho.
    """
    filas = []

    # Orden similar al export nativo observado: E ProbeCenter, H ProbeCenter,
    # H SensorMax, H TipCenter, H TipMax, Gradient ProbeCenter.
    d_probe = datos_por_location.get("probe-center")
    if d_probe:
        ts = d_probe.get("timestamp")
        fs = normalizar_frame_source(d_probe.get("source") or d_probe.get("frame_source"))
        freq = d_probe.get("frequency")
        ef = _campo_xyz_total(d_probe.get("e_field"))
        if ef is not None:
            filas.append(_fila_tidy(ts, ts_ref, fs, "IncidentElectricField",
                                    "ProbeCenter", 0, freq, *ef))

    for api_loc in API_LOCATIONS_EXPORT:
        d = datos_por_location.get(api_loc)
        if not d:
            continue
        ts = d.get("timestamp")
        fs = normalizar_frame_source(d.get("source") or d.get("frame_source"))
        freq = d.get("frequency")
        hf = _campo_xyz_total(d.get("h_field"))
        if hf is not None:
            filas.append(_fila_tidy(ts, ts_ref, fs, "IncidentMagneticField",
                                    LOCATION_API_TO_EXPORT.get(api_loc, api_loc),
                                    _extraer_sensor_index(d, api_loc), freq, *hf))

    if d_probe:
        ts = d_probe.get("timestamp")
        fs = normalizar_frame_source(d_probe.get("source") or d_probe.get("frame_source"))
        gr = _campo_xyz_total(d_probe.get("h_field_gradient"), total_keys=("total", "rms_total"))
        if gr is not None:
            filas.append(_fila_tidy(ts, ts_ref, fs, "IncidentMagneticFieldGradient",
                                    "ProbeCenter", 0, None, *gr))

    return filas

# ═══════════════════════════════════════════════════════════════
# HILO COLECTOR — recoge medidas a alta velocidad filtrando repetidas
# ═══════════════════════════════════════════════════════════════

class ColectorDatos:
    def __init__(self, api_location="sensor-max"):
        self.api_location         = api_location
        self.filas_measurement    = []       # Formato interno ancho usado para promedios/Excel
        self.filas_measurement_tidy = []     # Formato MAGPy 3.4 para Measurements.csv
        self.filas_compliance     = []
        self.ultimo_timestamp     = None
        self.ultimo_ts_compliance = None
        self.last_update_time     = None
        self.ts_inicio_ms         = None   # FIX: primer timestamp real para ventana relativa
        self.debug_api_measurements = []   # Primeras respuestas raw para diagnosticar campos extra
        self._tidy_seen           = set()
        self._ultimo_aviso_bloqueo = 0
        self._activo              = False
        self._hilo                = None
        self._hilo_compliance     = None   # FIX: hilo separado para no bloquear el polling
        self._hilo_tidy           = None   # FIX: hilo separado para no bloquear el polling de measurement

    def iniciar(self):
        self.filas_measurement    = []
        self.filas_measurement_tidy = []
        self.filas_compliance     = []
        self.ultimo_timestamp     = None
        self.ultimo_ts_compliance = None
        self.ts_inicio_ms         = None
        self.debug_api_measurements = []
        self._tidy_seen           = set()
        # Forzar redescubrimiento de endpoints en cada nueva sesión de medida
        ColectorDatos._endpoint_history_measurement = None
        ColectorDatos._endpoint_history_compliance  = None
        self._activo = True
        self._hilo             = threading.Thread(target=self._bucle,            daemon=True)
        self._hilo_compliance  = threading.Thread(target=self._bucle_compliance, daemon=True)
        self._hilo_tidy        = threading.Thread(target=self._bucle_tidy,       daemon=True)
        self._hilo.start()
        self._hilo_compliance.start()
        self._hilo_tidy.start()

    def detener(self):
        self._activo = False
        if self._hilo:
            self._hilo.join(timeout=5)
        if self._hilo_compliance:
            self._hilo_compliance.join(timeout=5)
        if self._hilo_tidy:
            self._hilo_tidy.join(timeout=5)

    def _bucle(self):
        sin_datos_consecutivos = 0
        while self._activo:
            hubo_dato = self._leer()
            if hubo_dato:
                sin_datos_consecutivos = 0
                # Dato recibido: volvemos a pedir inmediatamente
            else:
                sin_datos_consecutivos += 1
                if self.last_update_time and (time.time() - self.last_update_time > 5):
                    ahora = time.time()
                    if ahora - self._ultimo_aviso_bloqueo > 30:
                        msg("MAGPy no devuelve datos nuevos desde hace mas de 5 s", "AVISO")
                        self._ultimo_aviso_bloqueo = ahora
                # Sin dato nuevo: esperamos un poco antes de reintentar
                # para no saturar la CPU (50ms << 787ms de cadencia MAGpy)
                time.sleep(0.05)

    def _bucle_compliance(self):
        """Hilo separado para compliance — no bloquea el polling de measurement."""
        while self._activo:
            self._leer_compliance()
            time.sleep(POLL_INTERVAL_S * 4)   # Compliance no cambia tan rápido

    def _bucle_tidy(self):
        """
        Hilo separado y de baja frecuencia (1 Hz) para el export tidy MAGPy 3.4.
        FIX: antes se llamaba a _capturar_tidy_desde_api() dentro de _leer(), en el
        propio hilo de polling de measurement. Esa función hace hasta 6 peticiones
        HTTP extra (3 locations x 2 endpoints) por cada frame, lo que ralentizaba
        el polling y provocaba pérdida de frames en la ventana de promediado
        (afectando a los valores volcados al Excel). Se desacopla aquí para que
        el polling de measurement no pierda cadencia.
        """
        while self._activo:
            if self.filas_measurement:
                self._capturar_tidy_desde_api(self.filas_measurement[-1])
            time.sleep(1.0)

    # ── Descubrimiento de endpoint histórico (se hace una sola vez) ──────────
    _endpoint_history_measurement = None  # None = aún no descubierto
    _endpoint_history_compliance  = None

    @classmethod
    def _descubrir_endpoint_history(cls, tipo="measurement"):
        """
        Prueba los posibles endpoints históricos de la API MAGpy.
        Devuelve el nombre del endpoint que responde, o None si ninguno lo hace.
        tipo: "measurement" | "compliance"
        """
        candidatos = [
            f"data/history/{tipo}",
            f"data/all/{tipo}",
            f"data/{tipo}/history",
            f"data/{tipo}/list",
        ]
        for ep in candidatos:
            # Probamos con un since=0 para ver si el endpoint existe
            r = api_get(f"{ep}?since=0&location=sensor-max")
            if r is not None:
                msg(f"Endpoint histórico encontrado: /{ep}", "OK")
                return ep
        msg(f"No hay endpoint histórico para '{tipo}', usando solo /data/latest/{tipo}.", "AVISO")
        return False   # False = descubierto pero no disponible

    def _leer(self):
        """
        Solicita todos los frames nuevos desde el último timestamp.
        Devuelve True si se capturó al menos un frame nuevo, False si no.
        """
        capturados = 0
        try:
            if ColectorDatos._endpoint_history_measurement is None:
                ColectorDatos._endpoint_history_measurement = \
                    ColectorDatos._descubrir_endpoint_history("measurement")

            usar_history = ColectorDatos._endpoint_history_measurement

            if usar_history:
                since = self.ultimo_timestamp if self.ultimo_timestamp is not None else 0
                respuesta = api_get(
                    f"{usar_history}?since={since}&location={self.api_location}"
                )
                if isinstance(respuesta, list):
                    frames = respuesta
                elif isinstance(respuesta, dict):
                    frames = (respuesta.get("items")
                              or respuesta.get("data")
                              or respuesta.get("measurements")
                              or [])
                else:
                    frames = []
            else:
                datos = api_get(f"data/latest/measurement?location={self.api_location}")
                frames = [datos] if datos else []

            for datos in frames:
                if not datos:
                    continue
                ts = datos.get("timestamp")
                if ts is None or ts == self.ultimo_timestamp:
                    continue
                if any(f["timestamp"] == ts for f in self.filas_measurement[-10:]):
                    continue

                self.ultimo_timestamp = ts
                self.last_update_time = time.time()
                if self.ts_inicio_ms is None:
                    self.ts_inicio_ms = ts

                hf   = datos.get("h_field") or {}
                ef   = datos.get("e_field") or {}
                gr   = datos.get("h_field_gradient") or {}
                freq = datos.get("frequency", 0)

                self.filas_measurement.append({
                    "timestamp":            ts,
                    "frequency":            freq,
                    "h_field_rms_x":        hf.get("rms_x", 0),
                    "h_field_rms_y":        hf.get("rms_y", 0),
                    "h_field_rms_z":        hf.get("rms_z", 0),
                    "e_field_rms_x":        ef.get("rms_x", 0),
                    "e_field_rms_y":        ef.get("rms_y", 0),
                    "e_field_rms_z":        ef.get("rms_z", 0),
                    "h_field_gradient_x":   gr.get("x", 0),
                    "h_field_gradient_y":   gr.get("y", 0),
                    "h_field_gradient_z":   gr.get("z", 0),
                    "extrapolation_factor": datos.get("extrapolation_factor", 0),
                    "gradient_valid":       datos.get("h_field_gradient_valid", False),
                    "e_valid":              datos.get("e_field_valid", False),
                    "source":               datos.get("source", datos.get("frame_source", "")),
                })
                # FIX: _capturar_tidy_desde_api ya NO se llama aquí — se movió a
                # _bucle_tidy (hilo propio a 1 Hz) para no penalizar la cadencia
                # de este polling con las llamadas HTTP extra que hace.
                capturados += 1

        except Exception as e:
            msg(f"DEBUG _leer excepcion: {type(e).__name__}: {e}", "ERROR")

        return capturados > 0

    def _capturar_tidy_desde_api(self, datos_base):
        """Captura un snapshot en formato largo MAGPy 3.4 para todas las locations posibles."""
        datos_por_location = {}
        for loc in API_LOCATIONS_EXPORT:
            if loc == self.api_location:
                d = datos_base
            else:
                d = (api_get(f"data/current/measurement?location={loc}")
                     or api_get(f"data/latest/measurement?location={loc}"))
            if isinstance(d, dict) and d.get("timestamp") is not None:
                datos_por_location[loc] = d
                if len(self.debug_api_measurements) < 12:
                    self.debug_api_measurements.append({"location": loc, "data": d})

        if not datos_por_location:
            return 0

        filas = construir_filas_measurement_tidy(datos_por_location, self.ts_inicio_ms)
        nuevas = 0
        for f in filas:
            clave = (f.get("timestamp"), f.get("frame_source"), f.get("quantity"),
                     f.get("location"), f.get("sensor_index"), f.get("frequency"))
            if clave in self._tidy_seen:
                continue
            self._tidy_seen.add(clave)
            self.filas_measurement_tidy.append(f)
            nuevas += 1
        return nuevas

    def _leer_compliance(self):
        """Lee compliance con endpoint histórico si está disponible."""
        try:
            if ColectorDatos._endpoint_history_compliance is None:
                ColectorDatos._endpoint_history_compliance = \
                    ColectorDatos._descubrir_endpoint_history("compliance")

            usar_history = ColectorDatos._endpoint_history_compliance
            ts_ref = self.ultimo_ts_compliance or 0

            if usar_history:
                respuesta = api_get(
                    f"{usar_history}?since={ts_ref}&location={self.api_location}"
                )
                if isinstance(respuesta, list):
                    bloques = respuesta
                elif isinstance(respuesta, dict):
                    bloques = (respuesta.get("items")
                               or respuesta.get("data")
                               or [respuesta])
                else:
                    bloques = []
            else:
                comp = api_get(f"data/latest/compliance?location={self.api_location}")
                bloques = [comp] if comp else []

            for comp in bloques:
                if not comp:
                    continue
                ts = comp.get("timestamp") or self.ultimo_timestamp
                if ts is None or ts == self.ultimo_ts_compliance:
                    continue
                self.ultimo_ts_compliance = ts
                for item in comp.get("results", []):
                    ext = item.get("extrapolated") or {}
                    self.filas_compliance.append({
                        "timestamp":                  ts,
                        "standard":                   item.get("standard", ""),
                        "quantity":                   item.get("quantity", ""),
                        "unit":                       item.get("unit", ""),
                        "environment":                item.get("environment", ""),
                        "frequency":                  item.get("frequency", ""),
                        "value_rms":                  item.get("value_rms", ""),
                        "value_dB":                   item.get("value_dB", ""),
                        "safety_factor":              item.get("safety_factor", ""),
                        "extrapolation_factor":       item.get("extrapolation_factor", ""),
                        "conclusion":                 item.get("conclusion", ""),
                        "extrapolated_value_rms":     ext.get("value_rms", ""),
                        "extrapolated_value_dB":      ext.get("value_dB", ""),
                        "extrapolated_safety_factor": ext.get("safety_factor", ""),
                        "extrapolated_conclusion":    ext.get("conclusion", ""),
                        "enhancement_factor":         item.get("enhancement_factor", ""),
                    })
        except Exception:
            pass

    def n_filas(self):
        return len(self.filas_measurement)


# ═══════════════════════════════════════════════════════════════
# CAPTURA Y GUARDADO
# ═══════════════════════════════════════════════════════════════

def capturar_pantalla(carpeta, nombre_fase):
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = os.path.join(carpeta, f"screenshot_{nombre_fase.replace(' ','_')}_{ts}.png")
    pyautogui.screenshot().save(ruta)
    msg(f"Captura guardada: {os.path.basename(ruta)}", "OK")
    return ruta

def _csv_bytes(filas, campos):
    buf = io.StringIO()
    w   = csv.DictWriter(buf, fieldnames=campos, extrasaction="ignore")
    w.writeheader(); w.writerows(filas)
    return buf.getvalue().encode("utf-8")

def _construir_spectra_tidy(api_location):
    """Devuelve Spectra en formato largo MAGPy 3.4 para la location seleccionada."""
    sp  = api_get(f"data/latest/spectra?location={api_location}") or {}
    frq = sp.get("frequency") or []
    hf  = sp.get("h_field") or {}
    ef  = sp.get("e_field") or {}
    n   = len(frq)

    def _lista(d, *keys):
        for k in keys:
            v = d.get(k) if isinstance(d, dict) else None
            if isinstance(v, list):
                return v
        return [None] * n

    filas = []
    location = LOCATION_API_TO_EXPORT.get(api_location, api_location)
    for quantity, campo in (("IncidentMagneticField", hf), ("IncidentElectricField", ef)):
        xs = _lista(campo, "rms_x", "x")
        ys = _lista(campo, "rms_y", "y")
        zs = _lista(campo, "rms_z", "z")
        ts = _lista(campo, "rms_total", "total")
        mx = _lista(campo, "rms_max", "max")
        for i in range(n):
            x = xs[i] if i < len(xs) else None
            y = ys[i] if i < len(ys) else None
            z = zs[i] if i < len(zs) else None
            total = ts[i] if i < len(ts) else None
            if total is None:
                total = _vector_total(x, y, z)
            filas.append({
                "frequency": frq[i],
                "quantity": quantity,
                "location": location,
                "x": x,
                "y": y,
                "z": z,
                "total": total,
                "max": mx[i] if i < len(mx) else None,
            })
    return filas

def guardar_datos_zip(carpeta, nombre_fase, colector, api_location):
    msg(f"Empaquetando {colector.n_filas()} filas unicas medidas...", "ESPERA")
    ts_str     = datetime.now().strftime("%Y-%m-%dT%H_%M_%S")
    nombre_zip = f"Data_{nombre_fase.replace(' ','_')}_{ts_str}.zip"
    ruta_zip   = os.path.join(carpeta, nombre_zip)

    filas_measurements = list(colector.filas_measurement_tidy)
    if not filas_measurements:
        msg("No se capturaron filas en formato MAGPy 3.4. Se guardará Measurements con cabecera correcta pero sin datos.", "AVISO")

    filas_spec = _construir_spectra_tidy(api_location)

    metadata = {"timestamp": ts_str, "fase": nombre_fase,
                "n_muestras_unicas": colector.n_filas(),
                "n_measurements_tidy": len(filas_measurements),
                "location_utilizada_para_promedio": api_location,
                "measurements_format": "MAGPy 3.4 tidy API transform",
                "nota": "Measurements.csv se genera desde la API documentada, sin pulsar SAVE DATA.",
                "version": api_get("version") or {},
                "device":  api_get("device/info") or {}}

    campos_m = MEASUREMENTS_TIDY_COLUMNS
    campos_s = ["frequency", "quantity", "location", "x", "y", "z", "total", "max"]
    campos_c = ["timestamp","standard","quantity","unit","environment","frequency",
                "value_rms","value_dB","safety_factor","extrapolation_factor","conclusion",
                "extrapolated_value_rms","extrapolated_value_dB",
                "extrapolated_safety_factor","extrapolated_conclusion","enhancement_factor"]

    with zipfile.ZipFile(ruta_zip, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"Measurements_{ts_str}.csv",
                   _csv_bytes(filas_measurements, campos_m))
        z.writestr(f"Spectra_{ts_str}.csv",
                   _csv_bytes(filas_spec, campos_s))
        z.writestr(f"Compliance_{ts_str}.csv",
                   _csv_bytes(colector.filas_compliance, campos_c))
        z.writestr(f"Metadata_{ts_str}.json",
                   json.dumps(metadata, indent=2).encode("utf-8"))
        # Debug útil para comprobar si la API expone sensor_index/sensor_num con otro nombre.
        z.writestr(f"DEBUG_API_Measurements_raw_{ts_str}.json",
                   json.dumps(colector.debug_api_measurements, indent=2).encode("utf-8"))

    msg(f"ZIP guardado: {nombre_zip} ({len(filas_measurements)} filas Measurements formato MAGPy 3.4)", "OK")
    return ruta_zip


# ═══════════════════════════════════════════════════════════════
# PROMEDIADO — campo H y campo E
# ═══════════════════════════════════════════════════════════════

def calcular_h_total(x, y, z):
    try:
        return math.sqrt(float(x)**2 + float(y)**2 + float(z)**2)
    except (TypeError, ValueError):
        return None

def calcular_e_total(x, y, z):
    try:
        return math.sqrt(float(x)**2 + float(y)**2 + float(z)**2)
    except (TypeError, ValueError):
        return None

def promediar_medidas(filas_measurement, ms_ventana, ts_inicio_ms=None):
    """
    Calcula promedios de H y E sobre la ventana ms_ventana.
    ts_inicio_ms: timestamp del primer frame capturado por el colector.
    Si no se pasa, se usa el mínimo del DataFrame (comportamiento anterior).
    """
    if not filas_measurement:
        return {"h_prom": None, "e_prom": None, "freq_hz": None, "freq_mhz": None}

    df = pd.DataFrame(filas_measurement)
    ts_ref = ts_inicio_ms if ts_inicio_ms is not None else df["timestamp"].min()
    df["ts_rel"] = df["timestamp"] - ts_ref
    # Solo datos dentro de la ventana (y con ts_rel >= 0 para evitar frames previos)
    df_v = df[(df["ts_rel"] >= 0) & (df["ts_rel"] <= ms_ventana)].copy()

    if df_v.empty:
        msg("AVISO: ningún frame dentro de la ventana temporal, usando todos los datos.", "AVISO")
        df_v = df.copy()

    df_v["h_total"] = df_v.apply(
        lambda r: calcular_h_total(r.get("h_field_rms_x", 0), r.get("h_field_rms_y", 0), r.get("h_field_rms_z", 0)), axis=1)

    df_v["e_total"] = df_v.apply(
        lambda r: calcular_e_total(r.get("e_field_rms_x", 0), r.get("e_field_rms_y", 0), r.get("e_field_rms_z", 0)), axis=1)

    h_prom = df_v["h_total"].mean() if df_v["h_total"].notna().any() else None
    e_prom = df_v["e_total"].mean() if df_v["e_total"].notna().any() else None

    freq_hz  = df_v["frequency"].mode().iloc[0] if not df_v["frequency"].empty else None
    freq_mhz = round(freq_hz / 1_000_000, 6) if freq_hz else None

    msg(f"promediar_medidas: {len(df_v)} frames usados de {len(df)} totales "
        f"(ventana {ms_ventana/60000:.1f} min)", "OK")

    return {
        "h_prom":    round(h_prom, 6) if h_prom is not None else None,
        "e_prom":    round(e_prom, 6) if e_prom is not None else None,
        "freq_hz":   freq_hz,
        "freq_mhz":  freq_mhz,
    }


# ═══════════════════════════════════════════════════════════════
# ESCRITURA EN EXCEL
# ═══════════════════════════════════════════════════════════════

def encontrar_primera_fila_libre(ws, col_ref, first_data_row):
    r = first_data_row
    while ws.cell(r, col_ref).value is not None:
        r += 1
    return r

def escribir_fila_excel(ws, fila, cols, technology, wpt_client, battery,
                        nombre_posicion, distancia_cm, freq_mhz_e, e_prom,
                        fecha_e, freq_mhz_h, h_prom, fecha_h):
    ws.cell(fila, cols["technology"]).value = technology
    ws.cell(fila, cols["wpt"]).value        = wpt_client
    ws.cell(fila, cols["battery"]).value    = battery
    ws.cell(fila, cols["testside"]).value   = nombre_posicion
    ws.cell(fila, cols["distance"]).value   = distancia_cm

    if freq_mhz_e is not None:
        ws.cell(fila, cols["freq_e"]).value = freq_mhz_e
    if e_prom is not None:
        ws.cell(fila, cols["efield"]).value = round(e_prom, 4)
    ws.cell(fila, cols["date_e"]).value = fecha_e

    if freq_mhz_h is not None:
        ws.cell(fila, cols["freq_h"]).value = freq_mhz_h
    if h_prom is not None:
        ws.cell(fila, cols["hfield"]).value = round(h_prom, 4)
    ws.cell(fila, cols["date_h"]).value = fecha_h

def _abrir_excel_protegido(ruta_excel, keep_vba):
    """
    Descifra el .xlsm/.xlsx protegido con contraseña de Office en memoria
    (usando msoffcrypto-tool) y devuelve el Workbook de openpyxl ya abierto.
    Si el archivo no está cifrado, hace fallback a load_workbook directo.
    """
    global _EXCEL_PASSWORD
    with open(ruta_excel, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        if not office_file.is_encrypted():
            f.seek(0)
            return load_workbook(io.BytesIO(f.read()), keep_vba=keep_vba)

        if not _EXCEL_PASSWORD:
            _EXCEL_PASSWORD = getpass.getpass(
                "  Archivo Excel protegido. Contrasena (no se muestra en pantalla): "
            )
        buffer_descifrado = io.BytesIO()
        office_file.load_key(password=_EXCEL_PASSWORD)
        office_file.decrypt(buffer_descifrado)
        buffer_descifrado.seek(0)
        return load_workbook(buffer_descifrado, keep_vba=keep_vba)


def _guardar_excel_protegido(wb, ruta_excel):
    """
    Guarda el Workbook en un buffer y lo vuelve a cifrar con la misma
    contraseña de Office antes de escribirlo a disco, para no perder
    la protección del archivo original.
    """
    buffer_plano = io.BytesIO()
    wb.save(buffer_plano)
    buffer_plano.seek(0)

    if not _EXCEL_PASSWORD:
        # No debería ocurrir (ya se pidió al abrir), pero por seguridad
        # guardamos sin cifrar antes que perder los datos de la medida.
        with open(ruta_excel, "wb") as f:
            f.write(buffer_plano.read())
        msg("Excel guardado SIN proteger (no habia contrasena en sesion).", "AVISO")
        return

    office_file = msoffcrypto.OfficeFile(buffer_plano)
    buffer_cifrado = io.BytesIO()
    office_file.encrypt(password=_EXCEL_PASSWORD, outfile=buffer_cifrado)
    buffer_cifrado.seek(0)
    with open(ruta_excel, "wb") as f:
        f.write(buffer_cifrado.read())


def actualizar_excel(ruta_excel, region, nombre_posicion, distancia_cm,
                     promedios, technology, wpt_client, battery,
                     fecha_fin, frecuencia_khz):
    """
    Escritura en Excel: best-effort. Si el archivo esta protegido con
    contrasena, se pide por consola (input oculto) SOLO si hace falta.
    Si falla por password incorrecta, se reintenta una vez; si sigue
    fallando, se avisa y se continua la sesion de medida sin abortar
    (los datos de la medida no se pierden, solo no se vuelcan al Excel).
    """
    global _EXCEL_PASSWORD

    if not ruta_excel or not os.path.isfile(ruta_excel):
        msg("Ruta de Excel no valida, omitiendo escritura.", "AVISO")
        return

    keep_vba = ruta_excel.lower().endswith(".xlsm")
    wb = None
    for intento in range(2):   # 1 intento normal + 1 reintento si la password es incorrecta
        try:
            wb = _abrir_excel_protegido(ruta_excel, keep_vba)
            break
        except Exception as e:
            es_error_password = _EXCEL_PASSWORD is not None and (
                "password" in str(e).lower() or "hmac" in str(e).lower()
                or isinstance(e, ValueError)
            )
            if es_error_password and intento == 0:
                msg("Contrasena incorrecta o fallo al descifrar. Reintentando...", "AVISO")
                _EXCEL_PASSWORD = None
                continue
            msg(f"No se pudo abrir el Excel: {type(e).__name__}: {e}. "
                f"Se omite la escritura en Excel para esta posicion.", "ERROR")
            return

    if wb is None:
        msg("No se pudo abrir el Excel tras reintentar. Se omite la escritura.", "ERROR")
        return

    sheet_name = CONFIG_REGION[region]["sheet"]
    if sheet_name not in wb.sheetnames:
        msg(f"Pestaña '{sheet_name}' no encontrada en el Excel.", "ERROR")
        wb.close()
        return

    ws      = wb[sheet_name]
    bloques = EXCEL_COL_MAP[region]
    fecha_str = fecha_fin.strftime("%Y-%m-%d")

    h_prom   = promedios.get("h_prom")
    e_prom   = promedios.get("e_prom")
    freq_mhz = promedios.get("freq_mhz") or round(frecuencia_khz / 1000, 6)

    for cols in bloques:
        fila = encontrar_primera_fila_libre(ws, cols["technology"], cols["first_data_row"])
        escribir_fila_excel(
            ws, fila, cols, technology, wpt_client, battery,
            nombre_posicion, distancia_cm, freq_mhz, e_prom, fecha_str,
            freq_mhz, h_prom, fecha_str
        )
        msg(f"  Excel '{sheet_name}' bloque col {cols['technology']}: fila {fila} escrita", "OK")

    try:
        _guardar_excel_protegido(wb, ruta_excel)
        msg(f"Excel guardado: {os.path.basename(ruta_excel)}", "OK")
    except Exception as e:
        msg(f"Error al guardar Excel: {type(e).__name__}: {e}. "
            f"Los datos de esta medida NO se volcaron al Excel, "
            f"pero el CSV/ZIP de la posicion sigue intacto.", "ERROR")


# ═══════════════════════════════════════════════════════════════
# LÓGICA DE MEDIDA POR FASE
# ═══════════════════════════════════════════════════════════════

def click_clear_gui():
    w, h = pyautogui.size()
    pyautogui.click(int(w * 0.885), int(h * 0.155))
    time.sleep(0.8)

def realizar_fase(fase, nombre_posicion, carpeta_posicion, peak_rms_anterior, api_location):
    peak_rms_actual = fase["peak_rms"]

    if peak_rms_actual != peak_rms_anterior:
        # ... (mantén aquí tu lógica de cambio manual de configuración) ...
        enter("Presiona ENTER cuando hayas cambiado el modo en MAGpy3")

    msg(f"Preparando fase: {fase['etiqueta']}...", "ESPERA")
    click_clear_gui()
    time.sleep(2) # Aumentamos tiempo de espera tras el click

    colector = ColectorDatos(api_location=api_location)
    colector.iniciar()

    if not iniciar_medida_con_reintentos(f"{nombre_posicion} - {fase['etiqueta']}"):
        colector.detener()
        return False, None, None

    barra_progreso(fase["duracion_s"], fase["etiqueta"])
    fecha_fin = datetime.now()

    # FIX: primero parar la adquisición en la API, luego detener el hilo colector
    parar_medida()
    time.sleep(1)      # Dar margen a la API para procesar el stop
    colector.detener()
    time.sleep(0.5)

    # Captura de pantalla y guardado de datos
    capturar_pantalla(carpeta_posicion, fase["etiqueta"])
    guardar_datos_zip(carpeta_posicion, fase["etiqueta"], colector, api_location)

    # Calcular promedios del campo H y E sobre la ventana temporal de la fase
    promedios = promediar_medidas(colector.filas_measurement, fase["ms_ventana"],
                                   ts_inicio_ms=colector.ts_inicio_ms)
    msg(f"Promedios calculados -- H: {promedios.get('h_prom')} A/m | "
        f"E: {promedios.get('e_prom')} V/m | "
        f"Freq: {promedios.get('freq_mhz')} MHz", "OK")

    return True, promedios, fecha_fin

def calcular_plan_medida(regiones):
    """
    Dado un conjunto de regiones, calcula:
    - duracion_s: tiempo total que hay que medir (el máximo de todas las fases)
    - ms_ventana por región: cada región promedia sobre su ventana propia
    - si alguna región necesita Peak, se añade fase extra al final

    Devuelve una lista de 'tramos' ordenados:
      [{"duracion_s": X, "peak_rms": "RMS"/"Peak", "regiones_que_usan_este_tramo": [...]}]
    y un dict {region: ms_ventana_rms, region: ms_ventana_peak}
    """
    # Ventanas de promediado por región (en ms)
    ventanas = {}
    for r in regiones:
        cfg = CONFIG_REGION[r]
        ventanas[r] = {f["peak_rms"]: f["ms_ventana"] for f in cfg["fases"]}

    # Duración total = máximo de todas las fases RMS + Peak si las hay
    duracion_rms_s  = max(
        max(f["duracion_s"] for f in CONFIG_REGION[r]["fases"] if f["peak_rms"] == "RMS")
        for r in regiones
    )
    necesita_peak = any(
        any(f["peak_rms"] == "Peak" for f in CONFIG_REGION[r]["fases"])
        for r in regiones
    )
    duracion_peak_s = 0
    if necesita_peak:
        duracion_peak_s = max(
            (max((f["duracion_s"] for f in CONFIG_REGION[r]["fases"] if f["peak_rms"] == "Peak"), default=0)
             for r in regiones),
            default=0
        )

    return {
        "duracion_rms_s":  duracion_rms_s,
        "duracion_peak_s": duracion_peak_s,
        "necesita_peak":   necesita_peak,
        "ventanas":        ventanas,
    }


def procesar_posicion(nombre_posicion, carpeta_raiz, regiones,
                      distancia_cm, frecuencia_khz, ruta_excel,
                      technology, wpt_client, battery):
    nombre_carpeta   = nombre_posicion.replace(" ", "_").replace("/", "-")
    carpeta_posicion = os.path.join(carpeta_raiz, nombre_carpeta)
    os.makedirs(carpeta_posicion, exist_ok=True)
    msg(f"Carpeta creada: {carpeta_posicion}", "OK")

    if distancia_cm < 0.75:
        api_location = "tip-max"
    else:
        api_location = "sensor-max"

    plan = calcular_plan_medida(regiones)
    nombres_regiones = ", ".join(CONFIG_REGION[r]["nombre"] for r in regiones)
    fases_desc = []
    for r in regiones:
        fases_desc.append(f"  {r}: " + " + ".join(f["etiqueta"] for f in CONFIG_REGION[r]["fases"]))

    sep(f"POSICION: {nombre_posicion}")
    beep(2)
    cuadro(
        f"ACCION REQUERIDA -- {nombre_posicion}",
        [f"Coloca el dispositivo en la posicion: {nombre_posicion}",
         f"Distancia de medida: {distancia_cm} cm",
         f"Location automatico asignado: '{api_location}'",
         f"Frecuencia configurada: {frecuencia_khz} kHz",
         f"Regiones: {nombres_regiones}",
         f"Duracion RMS: {plan['duracion_rms_s']//60} min"
         + (f"  +  Peak: {plan['duracion_peak_s']//60} min" if plan["necesita_peak"] else ""),
         "", "Promediados:"] + fases_desc +
        ["", "Asegurate de que:",
         "  * La sonda MAGpy3 esta correctamente posicionada",
         "  * El dispositivo bajo prueba esta activo",
         "  * MAGpy3 muestra senal en tiempo real"]
    )
    enter("Presiona ENTER para INICIAR las medidas de esta posicion")

    # ── FASE RMS ─────────────────────────────────────────────────────────────
    colector_rms = ColectorDatos(api_location=api_location)
    colector_rms.iniciar()

    if not iniciar_medida_con_reintentos(f"{nombre_posicion} - RMS"):
        colector_rms.detener()
        return carpeta_posicion, None

    fase_rms_label = f"{plan['duracion_rms_s']//60} min RMS"
    barra_progreso(plan["duracion_rms_s"], fase_rms_label)
    fecha_fin_rms = datetime.now()

    parar_medida()
    time.sleep(1)
    colector_rms.detener()
    time.sleep(0.5)

    capturar_pantalla(carpeta_posicion, fase_rms_label)
    guardar_datos_zip(carpeta_posicion, fase_rms_label, colector_rms, api_location)

    # Calcular promedios RMS para cada región con su ventana propia
    promedios_por_region = {}
    for r in regiones:
        ms_ventana_rms = plan["ventanas"][r].get("RMS")
        if ms_ventana_rms:
            prom = promediar_medidas(colector_rms.filas_measurement, ms_ventana_rms,
                                     ts_inicio_ms=colector_rms.ts_inicio_ms)
            promedios_por_region[r] = {"RMS": prom}
            msg(f"[{r}] RMS ({ms_ventana_rms//60000} min) -- "
                f"H: {prom.get('h_prom')} A/m | E: {prom.get('e_prom')} V/m", "OK")

    # ── FASE PEAK (solo si alguna región la necesita) ─────────────────────────
    if plan["necesita_peak"]:
        regiones_peak = [r for r in regiones
                         if plan["ventanas"][r].get("Peak") is not None]
        beep(1)
        cuadro("CAMBIAR A MODO PEAK",
               ["Pon Peak/RMS = Peak en MAGpy3 --> Configure --> Settings"])
        enter("Presiona ENTER cuando hayas cambiado a Peak")

        click_clear_gui()
        time.sleep(2)
        colector_peak = ColectorDatos(api_location=api_location)
        colector_peak.iniciar()

        exito_peak = iniciar_medida_con_reintentos(f"{nombre_posicion} - Peak")
        if not exito_peak:
            colector_peak.detener()
            beep(1)
            cuadro("RESTABLECER MODO",
                   ["Vuelve a poner Peak/RMS = RMS en MAGpy3 --> Configure settings"])
            enter("Presiona ENTER cuando hayas vuelto a RMS")
            return carpeta_posicion, None

        fase_peak_label = f"{plan['duracion_peak_s']//60} min Peak"
        barra_progreso(plan["duracion_peak_s"], fase_peak_label)
        fecha_fin_peak = datetime.now()

        parar_medida()
        time.sleep(1)
        colector_peak.detener()
        time.sleep(0.5)

        capturar_pantalla(carpeta_posicion, fase_peak_label)
        guardar_datos_zip(carpeta_posicion, fase_peak_label, colector_peak, api_location)

        for r in regiones_peak:
            ms_ventana_peak = plan["ventanas"][r]["Peak"]
            prom = promediar_medidas(colector_peak.filas_measurement, ms_ventana_peak,
                                     ts_inicio_ms=colector_peak.ts_inicio_ms)
            if r not in promedios_por_region:
                promedios_por_region[r] = {}
            promedios_por_region[r]["Peak"] = prom
            msg(f"[{r}] Peak ({ms_ventana_peak//60000} min) -- "
                f"H: {prom.get('h_prom')} A/m | E: {prom.get('e_prom')} V/m", "OK")

        beep(1)
        cuadro("RESTABLECER MODO",
               ["Vuelve a poner Peak/RMS = RMS en MAGpy3 --> Configure settings"])
        enter("Presiona ENTER cuando hayas vuelto a RMS")

    # ── ESCRITURA EN EXCEL ────────────────────────────────────────────────────
    for r in regiones:
        prom_rms  = promedios_por_region.get(r, {}).get("RMS")
        prom_peak = promedios_por_region.get(r, {}).get("Peak")
        # Para escritura en Excel usamos el promedio RMS como principal
        # (ISED escribe el peak en un segundo bloque si lo hay)
        promedios_excel = prom_rms or {}
        if promedios_excel:
            actualizar_excel(
                ruta_excel, r, nombre_posicion, distancia_cm,
                promedios_excel, technology, wpt_client, battery,
                fecha_fin_rms, frecuencia_khz
            )
            msg(f"[{r}] Excel actualizado.", "OK")

    beep(3)
    msg(f"Posicion '{nombre_posicion}' completada", "OK")

    # Devolver promedios de la primera región como resumen CSV
    prom_resumen = (promedios_por_region.get(regiones[0], {}).get("RMS")
                    or next(iter(promedios_por_region.values()), {}).get("RMS"))
    return carpeta_posicion, prom_resumen


# ═══════════════════════════════════════════════════════════════
# RESUMEN CSV AL FINAL
# ═══════════════════════════════════════════════════════════════

def guardar_csv_resumen(resultados, archivo_csv):
    if not resultados:
        return
    df = pd.DataFrame(resultados)
    df.to_csv(archivo_csv, index=False, sep=";", decimal=",")
    msg(f"CSV de resumen guardado: {archivo_csv}", "OK")
    sep("RESUMEN FINAL")
    print(df.to_string(index=False))


# ═══════════════════════════════════════════════════════════════
# ENTRADA DE DATOS
# ═══════════════════════════════════════════════════════════════

def pedir_regiones(actual=None):
    """Permite seleccionar una o varias regiones (CE, FCC, ISED) en una sola sesión."""
    print("\n  Regiones disponibles (puedes seleccionar una o varias):")
    for k, v in CONFIG_REGION.items():
        fases_str = " + ".join(f["etiqueta"] for f in v["fases"])
        print(f"    [{k}]  {v['nombre']}  -->  {fases_str}")
    print("\n  Ejemplos de entrada: CE  /  FCC  /  CE FCC  /  CE FCC ISED")
    print("  Comandos: B = atras | R = reiniciar parametros | Q = cancelar")
    while True:
        prompt = _prompt_actual("\n  Regiones a medir: ", actual)
        entrada = input_config(prompt, permitir_vacio=(actual is not None))
        if entrada == "" and actual is not None:
            return actual
        entrada = entrada.upper()
        seleccionadas = [r.strip() for r in entrada.replace(",", " ").split() if r.strip()]
        invalidas = [r for r in seleccionadas if r not in CONFIG_REGION]
        if not seleccionadas:
            print("  X Introduce al menos una region.")
        elif invalidas:
            print(f"  X Opciones no validas: {', '.join(invalidas)}")
        else:
            # Eliminar duplicados manteniendo orden
            vistas = set()
            seleccionadas = [r for r in seleccionadas if not (r in vistas or vistas.add(r))]
            nombres = ", ".join(CONFIG_REGION[r]["nombre"] for r in seleccionadas)
            print(f"  OK Regiones seleccionadas: {nombres}")
            return seleccionadas


def pedir_carpeta(prompt, actual=None):
    while True:
        ruta = input_config(_prompt_actual(f"  {prompt}: ", actual), permitir_vacio=(actual is not None))
        if ruta == "" and actual is not None:
            return actual
        try:
            os.makedirs(ruta, exist_ok=True)
            print(f"  OK Carpeta lista: {ruta}")
            return ruta
        except Exception as e:
            print(f"  X Error: {e}")


def pedir_archivo_excel(actual=None):
    print("\n  Ruta al archivo Excel de resultados (.xlsm o .xlsx)")
    print(r"  Ejemplo: C:\Users\MAGPy\Desktop\FAN37_07_Field_testing.xlsm")
    print("  Deja en blanco para omitir la escritura en Excel")
    ruta = input_config(_prompt_actual("  Archivo Excel: ", actual), permitir_vacio=True)
    if ruta == "" and actual is not None:
        return actual
    if not ruta:
        return None
    if not os.path.isfile(ruta):
        print(f"  ! Archivo no encontrado: {ruta}. Se omitira la escritura en Excel.")
        return None
    return ruta


# Se cambia de pedir_entero a pedir_numero para permitir introducir distancias como 0.5 cm
def pedir_numero(texto, mn=0.0, mx=999.0, actual=None):
    while True:
        valor = input_config(_prompt_actual(f"  {texto}: ", actual), permitir_vacio=(actual is not None))
        if valor == "" and actual is not None:
            return actual
        try:
            v = float(valor)
            if mn <= v <= mx:
                return v
            print(f"  X Valor entre {mn} y {mx}.")
        except ValueError:
            print("  X Introduce un numero valido, o escribe B para atras / R para reiniciar / Q para cancelar.")


def pedir_entero(texto, mn=0, mx=999, actual=None):
    while True:
        valor = input_config(_prompt_actual(f"  {texto}: ", actual), permitir_vacio=(actual is not None))
        if valor == "" and actual is not None:
            return actual
        try:
            v = int(valor)
            if mn <= v <= mx:
                return v
            print(f"  X Valor entre {mn} y {mx}.")
        except ValueError:
            print("  X Introduce un numero entero, o escribe B para atras / R para reiniciar / Q para cancelar.")


def pedir_frecuencia(actual=None):
    print("\n  Frecuencia de busqueda de pico (kHz). Rango: 3-10000 kHz")
    while True:
        valor = input_config(_prompt_actual("  Frecuencia (kHz): ", actual), permitir_vacio=(actual is not None))
        if valor == "" and actual is not None:
            return actual
        try:
            f = float(valor)
            if 3 <= f <= 10000:
                return f
            print("  X Fuera del rango 3-10000 kHz.")
        except ValueError:
            print("  X Introduce un numero, o escribe B para atras / R para reiniciar / Q para cancelar.")


def pedir_texto(prompt, default=None, actual=None):
    valor_actual = actual if actual is not None else None
    v = input_config(_prompt_actual(f"  {prompt}: ", valor_actual), permitir_vacio=True)
    if v == "" and actual is not None:
        return actual
    return v if v else default


def pedir_nombre_posicion(num, total):
    print(f"\n  Posicion {num} de {total}")
    nombre = input("  Nombre de esta posicion (ej: Front-Face, Right, 0cm): ").strip()
    return nombre if nombre else f"Posicion_{num}"

def mostrar_config_manual(cfg):
    cuadro(
        "CONFIGURA MANUALMENTE MAGpy3 ANTES DE CONTINUAR",
        ["Ve a MAGpy3 --> Configure settings (rueda dentada) y ajusta:",
         "", "  [Compliance]",
         f"    Standard            -->  {cfg['standard']}",
         f"    Environment         -->  {cfg['environment']}",
         f"    Health effect       -->  {cfg['health_effect']}",
         "", "  [Measurement]",
         f"    Peak/RMS            -->  {cfg['peak_rms_inicial']}",
         "    * Compliance location NO ES NECESARIO, LO GESTIONA LA API AUTOMATICAMENTE",
         "", "Cierra Settings cuando hayas terminado."]
    )
    enter("Presiona ENTER cuando hayas configurado MAGpy3")


def imprimir_resumen_sesion(regiones, carpeta_raiz, ruta_excel, n_posiciones,
                            distancia_cm, frecuencia, technology, wpt_client,
                            battery, archivo_csv, plan, nombres_regiones):
    duracion_total_min = (plan["duracion_rms_s"] + plan["duracion_peak_s"]) // 60

    sep("RESUMEN DE LA SESION")
    print(f"""
  Regiones          : {nombres_regiones}
  Posiciones        : {n_posiciones}
  Distancia         : {distancia_cm} cm
  Frecuencia        : {frecuencia} kHz
  Duracion/posicion : {duracion_total_min} min  (RMS: {plan['duracion_rms_s']//60} min""" +
          (f"  +  Peak: {plan['duracion_peak_s']//60} min)" if plan["necesita_peak"] else ")") + f"""
  Technology        : {technology}
  WPT Client        : {wpt_client}
  Battery           : {battery}
  Carpeta raiz      : {carpeta_raiz}
  Excel             : {ruta_excel or 'No configurado'}
  CSV salida        : {archivo_csv}

  Promediados por region:""")
    for r in regiones:
        for f in CONFIG_REGION[r]["fases"]:
            print(f"    [{r}]  {f['etiqueta']}  ({f['ms_ventana']//60000} min)")


def confirmar_configuracion():
    print("\n  Revisa el resumen antes de empezar.")
    print("  ENTER  --> continuar")
    print("  B      --> volver al parametro anterior")
    print("  R      --> reiniciar parametros")
    print("  Q      --> cancelar")
    while True:
        accion = input("  Opcion: ").strip().upper()
        if accion == "":
            return
        if accion == "B":
            raise VolverAtrasConfiguracion
        if accion == "R":
            raise ReiniciarConfiguracion
        if accion == "Q":
            raise CancelarSesion
        print("  X Opcion no valida. Usa ENTER, B, R o Q.")


def _valores_configuracion_iniciales():
    return {
        "regiones": None,
        "carpeta_raiz": None,
        "ruta_excel": None,
        "n_posiciones": None,
        "distancia_cm": None,
        "frecuencia": None,
        "technology": None,
        "wpt_client": None,
        "battery": None,
    }


def pedir_configuracion_sesion():
    """Pide parámetros iniciales con opción de atrás, reinicio o cancelación."""
    valores = _valores_configuracion_iniciales()

    pasos = [
        ("regiones",     lambda: pedir_regiones(valores["regiones"])),
        ("carpeta_raiz", lambda: pedir_carpeta("Carpeta donde se guardaran los resultados", valores["carpeta_raiz"])),
        ("ruta_excel",   lambda: pedir_archivo_excel(valores["ruta_excel"])),
        ("n_posiciones", lambda: pedir_entero("Numero de posiciones a medir", 1, 50, valores["n_posiciones"])),
        ("distancia_cm", lambda: pedir_numero("Distancia de medida (cm)", 0.0, 200.0, valores["distancia_cm"])),
        ("frecuencia",   lambda: pedir_frecuencia(valores["frecuencia"])),
        ("technology",   lambda: pedir_texto("Technology (ej: NFC, WPT, BT)", "N/A", valores["technology"])),
        ("wpt_client",   lambda: pedir_texto("WPT Client (N/A si no es WPT)", "N/A", valores["wpt_client"])),
        ("battery",      lambda: pedir_texto("Battery level (N/A si no aplica)", "N/A", valores["battery"])),
    ]

    indice = 0
    primera_pantalla = True

    while True:
        try:
            if primera_pantalla:
                sep("CONFIGURACION DE LA SESION")
                mostrar_ayuda_parametros()
                primera_pantalla = False

            if indice < len(pasos):
                if indice == 6:
                    print("\n  Datos para el Excel (deja en blanco para N/A):")
                clave, funcion_paso = pasos[indice]
                valores[clave] = funcion_paso()
                indice += 1
                continue

            ts_sesion   = datetime.now().strftime("%Y%m%d_%H%M%S")
            archivo_csv = os.path.join(valores["carpeta_raiz"], f"promedios_{ts_sesion}.csv")
            plan = calcular_plan_medida(valores["regiones"])
            nombres_regiones = ", ".join(CONFIG_REGION[r]["nombre"] for r in valores["regiones"])

            imprimir_resumen_sesion(
                valores["regiones"], valores["carpeta_raiz"], valores["ruta_excel"], valores["n_posiciones"],
                valores["distancia_cm"], valores["frecuencia"], valores["technology"], valores["wpt_client"],
                valores["battery"], archivo_csv, plan, nombres_regiones
            )
            confirmar_configuracion()

            return {
                "regiones": valores["regiones"],
                "carpeta_raiz": valores["carpeta_raiz"],
                "ruta_excel": valores["ruta_excel"],
                "n_posiciones": valores["n_posiciones"],
                "distancia_cm": valores["distancia_cm"],
                "frecuencia": valores["frecuencia"],
                "technology": valores["technology"],
                "wpt_client": valores["wpt_client"],
                "battery": valores["battery"],
                "archivo_csv": archivo_csv,
                "plan": plan,
                "nombres_regiones": nombres_regiones,
            }

        except VolverAtrasConfiguracion:
            if indice <= 0:
                msg("Ya estas en el primer parametro; no hay paso anterior.", "AVISO")
                indice = 0
            else:
                indice -= 1
                msg("Volviendo al parametro anterior.", "AVISO")
            continue
        except ReiniciarConfiguracion:
            valores = _valores_configuracion_iniciales()
            indice = 0
            primera_pantalla = True
            msg("Reiniciando seleccion de parametros.", "AVISO")
            continue
        except CancelarSesion:
            msg("Sesion cancelada antes de iniciar las medidas.", "ERROR")
            sys.exit(0)


# ═══════════════════════════════════════════════════════════════
# PROGRAMA PRINCIPAL
# ═══════════════════════════════════════════════════════════════

def main():
    sep("AUTOMATIZACION DE MEDIDAS MAGpy3 + PROMEDIADO + EXCEL")

    if not verificar_conexion():
        msg("Abre MAGpy3, verifica que el backend esta activo.", "ERROR")
        sys.exit(1)

    config = pedir_configuracion_sesion()
    regiones         = config["regiones"]
    carpeta_raiz     = config["carpeta_raiz"]
    ruta_excel       = config["ruta_excel"]
    n_posiciones     = config["n_posiciones"]
    distancia_cm     = config["distancia_cm"]
    frecuencia       = config["frecuencia"]
    technology       = config["technology"]
    wpt_client       = config["wpt_client"]
    battery          = config["battery"]
    archivo_csv      = config["archivo_csv"]
    plan             = config["plan"]
    nombres_regiones = config["nombres_regiones"]

    if not validar_configuracion_magpy(regiones[0]):
        print("\n  X Corrige la configuracion en MAGpy3 y vuelve a ejecutar el script.")
        sys.exit(1)

    sep("CONFIGURANDO ADQUISICION VIA API")
    configurar_adquisicion(frecuencia)
    time.sleep(1)

    sep("INICIO DE MEDIDAS")
    t_inicio  = datetime.now()
    resultados = []

    for i in range(1, n_posiciones + 1):
        nombre_pos = pedir_nombre_posicion(i, n_posiciones)
        while True:
            carpeta_pos, promedios = procesar_posicion(
                nombre_pos, carpeta_raiz, regiones,
                distancia_cm, frecuencia, ruta_excel,
                technology, wpt_client, battery
            )
            if promedios:
                resultados.append({
                    "posicion":       nombre_pos,
                    "regiones":       "+".join(regiones),
                    "distancia_cm":   distancia_cm,
                    "frecuencia_khz": frecuencia,
                    "freq_mhz":       promedios.get("freq_mhz"),
                    "h_prom_A_m":     promedios.get("h_prom"),
                    "e_prom_V_m":     promedios.get("e_prom"),
                })
                break

            beep(2)
            cuadro(
                "POSICION NO COMPLETADA",
                [f"La posicion '{nombre_pos}' no se ha completado.",
                 "No se pasara a la siguiente posicion automaticamente.",
                 "",
                 "ENTER  --> repetir la misma posicion",
                 "A      --> abortar la sesion"]
            )
            accion = input("  Opcion: ").strip().upper()
            if accion == "A":
                msg("Sesion abortada. No se continuara con la siguiente posicion.", "ERROR")
                guardar_csv_resumen(resultados, archivo_csv)
                sys.exit(1)

    guardar_csv_resumen(resultados, archivo_csv)

    t_fin = datetime.now()
    dur   = t_fin - t_inicio
    h, m  = int(dur.total_seconds()//3600), int((dur.total_seconds()%3600)//60)

    beep(5)
    cuadro("SESION COMPLETADA",
           [f"Posiciones medidas : {n_posiciones}",
            f"Regiones           : {nombres_regiones}",
            f"Frecuencia         : {frecuencia} kHz",
            f"Hora inicio        : {t_inicio.strftime('%H:%M:%S')}",
            f"Hora fin           : {t_fin.strftime('%H:%M:%S')}",
            f"Duracion total     : {h}h {m}min",
            "",
            "Archivos generados:",
            f"  Carpeta raiz  -->  {carpeta_raiz}",
            f"  CSV resumen   -->  {archivo_csv}",
            f"  Excel         -->  {ruta_excel or 'No configurado'}"])

    return


if __name__ == "__main__":
    main()